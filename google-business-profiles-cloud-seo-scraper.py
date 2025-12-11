import os
import time
import json
import logging
import re
from typing import Optional, Dict, List, Tuple
from urllib.parse import urlparse, urljoin

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

for _name in ("urllib3", "urllib3.connection", "urllib3.connectionpool", "urllib3.util.ssl_"):
    logging.getLogger(_name).setLevel(logging.ERROR)

GOOGLE_CLOUD_API_KEY = os.getenv("GOOGLE_CLOUD_API_KEY")
GOOGLE_PROGRAMMABLE_SEARCH_ENGINE_CX = os.getenv("GOOGLE_PROGRAMMABLE_SEARCH_ENGINE_CX")


def env_flag(name: str, default: bool = True) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    value = value.strip().lower()
    if value in ("1", "true", "t", "yes", "y", "sim", "s"):
        return True
    if value in ("0", "false", "f", "no", "n", "nao", "não"):
        return False
    return default


USE_PLACES_API = env_flag("USE_PLACES_API", True)
USE_PAGESPEED_INSIGHTS_API = env_flag("USE_PAGESPEED_INSIGHTS_API", True)
USE_PROGRAMMABLE_SEARCH_ENGINE_API = env_flag("USE_PROGRAMMABLE_SEARCH_ENGINE_API", True)
USE_KNOWLEDGE_GRAPH_API = env_flag("USE_KNOWLEDGE_GRAPH_API", True)

BASE_QUERY = "local search"
CITIES = [
    "City, ST",
]

DEFAULT_COUNTRY_CODE = "55"
OUTPUT_XLSX = "google-business-profiles.xlsx"
KNOWN_PLACES_FILE = "known_places.json"
PAGESPEED_CACHE_FILE = "pagespeed_cache.json"
CSE_SITE_CACHE_FILE = "cse_site_cache.json"
CSE_BRAND_CACHE_FILE = "cse_brand_cache.json"
KG_CACHE_FILE = "kg_cache.json"

MAX_PAGESPEED_REQUESTS = 200
MAX_CSE_REQUESTS = 200
MAX_KG_REQUESTS = 200

SOCIAL_DOMAINS = [
    "facebook.com",
    "instagram.com",
    "linkedin.com",
    "youtube.com",
    "tiktok.com",
    "kwai.com",
    "twitter.com",
    "x.com",
    "pinterest.com",
    "whatsapp.com",
    "wa.me",
    "messenger.com",
    "t.me",
    "telegram.me",
]

DELIVERY_DOMAINS = [
    "ifood.com",
    "rappi.com",
    "ubereats.com",
    "99food.com",
    "deliverymuch.com",
    "aiqfome.com",
    "loggi.com",
]

pagespeed_cache: Dict[str, Dict] = {}
cse_site_cache: Dict[str, Dict] = {}
cse_brand_cache: Dict[str, Dict] = {}
kg_cache: Dict[str, Dict] = {}
pagespeed_requests = 0
cse_requests = 0
kg_requests = 0


def google_places_text_search(query: str, api_key: str, page_token: Optional[str] = None) -> Dict:
    url = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": "places.id,nextPageToken",
    }
    body: Dict[str, object] = {
        "textQuery": query,
        "languageCode": "pt-BR",
        "regionCode": "BR",
        "pageSize": 20,
    }
    if page_token:
        body["pageToken"] = page_token
    response = requests.post(url, headers=headers, json=body, timeout=20)
    response.raise_for_status()
    return response.json()


def google_place_details(place_id: str, api_key: str) -> Dict:
    url = f"https://places.googleapis.com/v1/places/{place_id}"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": ",".join(
            [
                "id",
                "displayName",
                "formattedAddress",
                "websiteUri",
                "internationalPhoneNumber",
                "nationalPhoneNumber",
                "rating",
                "userRatingCount",
                "types",
                "businessStatus",
                "googleMapsUri",
            ]
        ),
    }
    params = {
        "languageCode": "pt-BR",
        "regionCode": "BR",
    }
    response = requests.get(url, headers=headers, params=params, timeout=20)
    response.raise_for_status()
    return response.json()


def extract_city_state(address: Optional[str], search_city: str) -> Tuple[str, str]:
    if not address:
        city = search_city
        state = ""
        if "," in search_city:
            parts = [part.strip() for part in search_city.split(",") if part.strip()]
            if len(parts) >= 2:
                city = parts[0]
                state = parts[1]
        return city, state

    parts = [part.strip() for part in address.split(",") if part.strip()]
    city = ""
    state = ""

    for part in reversed(parts):
        if " - " in part:
            left, right = part.split(" - ", 1)
            city = left.strip()
            right = right.strip()
            match = re.search(r"\b([A-Z]{2})\b", right)
            if match:
                state = match.group(1)
            else:
                state = right
            break

    if not city:
        if len(parts) >= 2:
            city = parts[-2]
        else:
            city = search_city

    if not state:
        if "," in search_city:
            search_city_parts = [p.strip() for p in search_city.split(",") if p.strip()]
            if len(search_city_parts) >= 2:
                state = search_city_parts[1]

    return city, state


def normalize_phone_to_whatsapp_link(phone: Optional[str], country_code: str = DEFAULT_COUNTRY_CODE) -> Optional[str]:
    if not phone:
        return None
    digits = re.sub(r"\D", "", phone)
    if not digits:
        return None
    if not digits.startswith(country_code):
        digits = country_code + digits
    return f"https://wa.me/{digits}"


def classify_website(url: Optional[str]) -> Tuple[str, bool]:
    if not url:
        return "Desconhecido", False
    try:
        parsed = urlparse(url)
    except Exception:
        return "Desconhecido", False

    host = parsed.netloc.lower()
    if not host:
        return "Desconhecido", False
    if host.startswith("www."):
        host = host[4:]

    for domain in SOCIAL_DOMAINS:
        if host == domain or host.endswith("." + domain):
            return "Rede social", False

    for domain in DELIVERY_DOMAINS:
        if host == domain or host.endswith("." + domain):
            return "App delivery", False

    return "Site próprio", True


def site_scrape(url: str) -> Dict:
    result = {
        "https_ok": False,
        "has_whatsapp": False,
        "whatsapp_url": None,
        "tem_blog": False,
        "blog_url": None,
        "tem_amp": False,
        "amp_url": None,
        "facebook": None,
        "instagram": None,
        "youtube": None,
        "tiktok": None,
        "linkedin": None,
        "kwai": None,
        "messenger": None,
        "telegram": None,
        "pinterest": None,
        "x_twitter": None,
        "site_error": None,
    }

    try:
        response = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
        final_url = response.url or url
        if final_url.lower().startswith("https://"):
            result["https_ok"] = True
        html = response.text
    except Exception as exception:
        message = f"{type(exception).__name__}: {str(exception)}"
        if len(message) > 180:
            message = message[:177] + "..."
        result["site_error"] = message
        return result

    lower_html = html.lower()

    whatsapp_pattern = re.compile(r'(https?://(?:wa\.me|api\.whatsapp\.com)[^"\'\s<]+)', re.IGNORECASE)
    match_whatsapp = whatsapp_pattern.search(html)
    if match_whatsapp:
        result["has_whatsapp"] = True
        result["whatsapp_url"] = match_whatsapp.group(1).strip()
    elif "whatsapp" in lower_html:
        result["has_whatsapp"] = True

    blog_keywords = ["blog", "noticia", "notícias", "noticias", "news", "artigo", "artigos"]
    blog_link = None
    link_pattern = re.compile(r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', re.IGNORECASE | re.DOTALL)

    for href, anchor_text in link_pattern.findall(html):
        href_lower = (href or "").lower()
        anchor_lower = (anchor_text or "").lower()
        if any(keyword in href_lower for keyword in blog_keywords) or any(keyword in anchor_lower for keyword in blog_keywords):
            blog_link = href.strip()
            break

    try:
        final_url = response.url or url
    except Exception:
        final_url = url

    if blog_link:
        result["blog_url"] = urljoin(final_url, blog_link)
        result["tem_blog"] = True
    elif any(keyword in lower_html for keyword in ["/blog", "blog.", "/noticias", "/notícias", "/news", "/artigos"]):
        result["tem_blog"] = True

    amp_pattern = re.compile(r'<link[^>]+rel=["\']amphtml["\'][^>]+href=["\']([^"\']+)["\']', re.IGNORECASE)
    match_amp = amp_pattern.search(html)
    if match_amp:
        result["tem_amp"] = True
        result["amp_url"] = urljoin(final_url, match_amp.group(1).strip())

    social_patterns = {
        "facebook": re.compile(r'(https?://(?:www\.)?facebook\.com/[^"\'\s<]+)', re.IGNORECASE),
        "instagram": re.compile(r'(https?://(?:www\.)?instagram\.com/[^"\'\s<]+)', re.IGNORECASE),
        "youtube": re.compile(r'(https?://(?:www\.)?(?:youtube\.com|youtu\.be)/[^"\'\s<]+)', re.IGNORECASE),
        "tiktok": re.compile(r'(https?://(?:www\.)?tiktok\.com/[^"\'\s<]+)', re.IGNORECASE),
        "linkedin": re.compile(r'(https?://(?:www\.)?linkedin\.com/[^"\'\s<]+)', re.IGNORECASE),
        "kwai": re.compile(r'(https?://(?:www\.)?kwai\.com/[^"\'\s<]+)', re.IGNORECASE),
        "messenger": re.compile(r'(https?://(?:www\.)?messenger\.com/[^"\'\s<]+)', re.IGNORECASE),
        "telegram": re.compile(r'(https?://(?:t\.me|telegram\.me)/[^"\'\s<]+)', re.IGNORECASE),
        "pinterest": re.compile(r'(https?://(?:www\.)?pinterest\.com/[^"\'\s<]+)', re.IGNORECASE),
        "x_twitter": re.compile(r'(https?://(?:www\.)?(?:twitter\.com|x\.com)/[^"\'\s<]+)', re.IGNORECASE),
    }

    for key, pattern in social_patterns.items():
        match_social = pattern.search(html)
        if match_social:
            result[key] = match_social.group(1).strip()

    return result


def load_json_cache(path: str) -> Dict:
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as file:
            data = json.load(file)
        if isinstance(data, dict):
            return data
        return {}
    except Exception:
        return {}


def save_json_cache(path: str, data: Dict) -> None:
    try:
        with open(path, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)
    except Exception as exception:
        logging.warning(f"Falha ao salvar cache {path}: {exception}")


def load_known_places() -> set:
    if not os.path.exists(KNOWN_PLACES_FILE):
        return set()
    try:
        with open(KNOWN_PLACES_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)
        if isinstance(data, list):
            return set(data)
        return set()
    except Exception:
        return set()


def save_known_places(known: set) -> None:
    try:
        with open(KNOWN_PLACES_FILE, "w", encoding="utf-8") as file:
            json.dump(sorted(known), file, ensure_ascii=False, indent=2)
    except Exception as exception:
        logging.warning(f"Falha ao salvar {KNOWN_PLACES_FILE}: {exception}")


def sanitize_sheet_title(title: str) -> str:
    invalid_pattern = r'[\[\]\*/:\?]'
    cleaned = re.sub(invalid_pattern, " ", title).strip()
    if not cleaned:
        cleaned = "Dados"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned


def get_domain_root(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    try:
        parsed = urlparse(url)
    except Exception:
        return None

    host = parsed.netloc.lower()
    if not host:
        return None
    if host.startswith("www."):
        host = host[4:]
    return host


def get_pagespeed_for_domain(domain: str, site_url: str) -> Optional[Dict]:
    global pagespeed_requests

    if not GOOGLE_CLOUD_API_KEY or not USE_PAGESPEED_INSIGHTS_API:
        return None
    if domain in pagespeed_cache:
        return pagespeed_cache[domain]
    if pagespeed_requests >= MAX_PAGESPEED_REQUESTS:
        return None

    parsed = urlparse(site_url)
    scheme = "https" if parsed.scheme == "https" else "http"
    base_url = f"{scheme}://{parsed.netloc or domain}/"

    api_url = "https://www.googleapis.com/pagespeedonline/v5/runPagespeed"
    params = {
        "url": base_url,
        "strategy": "mobile",
        "category": ["performance", "seo"],
        "key": GOOGLE_CLOUD_API_KEY,
    }

    try:
        response = requests.get(api_url, params=params, timeout=60)
        if response.status_code != 200:
            return None
        data = response.json()
    except Exception:
        return None

    result: Dict[str, Optional[float]] = {}

    try:
        lighthouse = data.get("lighthouseResult", {})
        categories = lighthouse.get("categories", {})
        performance_score = categories.get("performance", {}).get("score")
        seo_category_score = categories.get("seo", {}).get("score") if "seo" in categories else None

        if performance_score is not None:
            result["ps_performance"] = round(float(performance_score) * 100)
        else:
            result["ps_performance"] = None

        if seo_category_score is not None:
            result["ps_seo"] = round(float(seo_category_score) * 100)
        else:
            result["ps_seo"] = None

        audits = lighthouse.get("audits", {})
        lcp_lab = audits.get("largest-contentful-paint", {}).get("numericValue")
        cls_lab = audits.get("cumulative-layout-shift", {}).get("numericValue")
        inp_lab = audits.get("interactive", {}).get("numericValue")

        loading = data.get("loadingExperience", {})
        metrics = loading.get("metrics", {})
        lcp_field = metrics.get("LARGEST_CONTENTFUL_PAINT_MS", {}).get("percentile")
        cls_field = metrics.get("CUMULATIVE_LAYOUT_SHIFT_SCORE", {}).get("percentile")
        inp_field = metrics.get("INTERACTION_TO_NEXT_PAINT", {}).get("percentile")

        if lcp_field is not None:
            result["lcp_ms"] = lcp_field
        elif lcp_lab is not None:
            result["lcp_ms"] = lcp_lab
        else:
            result["lcp_ms"] = None

        if cls_field is not None:
            result["cls"] = cls_field / 100 if isinstance(cls_field, (int, float)) else None
        elif cls_lab is not None:
            result["cls"] = cls_lab
        else:
            result["cls"] = None

        if inp_field is not None:
            result["inp_ms"] = inp_field
        elif inp_lab is not None:
            result["inp_ms"] = inp_lab
        else:
            result["inp_ms"] = None

        result["experience_category"] = loading.get("overall_category")
    except Exception:
        pass

    pagespeed_cache[domain] = result
    pagespeed_requests += 1
    return result


def custom_search(query: str) -> Optional[Dict]:
    global cse_requests

    if (
        not GOOGLE_CLOUD_API_KEY
        or not GOOGLE_PROGRAMMABLE_SEARCH_ENGINE_CX
        or not USE_PROGRAMMABLE_SEARCH_ENGINE_API
    ):
        return None

    if cse_requests >= MAX_CSE_REQUESTS:
        return None

    url = "https://www.googleapis.com/customsearch/v1"
    params = {
        "key": GOOGLE_CLOUD_API_KEY,
        "cx": GOOGLE_PROGRAMMABLE_SEARCH_ENGINE_CX,
        "q": query,
        "num": 10,
    }

    try:
        response = requests.get(url, params=params, timeout=20)
        if response.status_code != 200:
            return None
        data = response.json()
    except Exception:
        return None

    cse_requests += 1
    return data


def parse_total_results(value) -> Optional[int]:
    if value is None:
        return None
    as_string = str(value)
    digits = re.sub(r"\D", "", as_string)
    if not digits:
        return None
    try:
        return int(digits)
    except Exception:
        return None


def get_site_index_info(domain: str) -> Optional[Dict]:
    if not domain:
        return None
    if domain in cse_site_cache:
        return cse_site_cache[domain]

    data = custom_search(f"site:{domain}")
    if not data:
        return None

    info: Dict[str, Optional[int]] = {}

    try:
        total = data.get("searchInformation", {}).get("totalResults")
        info["total_results"] = parse_total_results(total)
    except Exception:
        info["total_results"] = None

    cse_site_cache[domain] = info
    return info


def get_brand_search_info(domain: Optional[str], site_url: Optional[str], name: str, city: str) -> Optional[Dict]:
    if not name:
        return None

    cache_key = f"{name}|{city}|{domain or ''}"
    if cache_key in cse_brand_cache:
        return cse_brand_cache[cache_key]

    data = custom_search(f'"{name}" {city}')
    if not data:
        return None

    position = None
    site_root = None

    if site_url:
        site_root = get_domain_root(site_url)
    if not site_root and domain:
        site_root = domain

    items = data.get("items", [])

    if items and site_root:
        for index, item in enumerate(items, start=1):
            link = item.get("link", "") or ""
            display_link = item.get("displayLink", "") or ""

            try:
                link_host = urlparse(link).netloc.lower()
            except Exception:
                link_host = ""

            display_host = display_link.lower()

            if link_host.startswith("www."):
                link_host = link_host[4:]
            if display_host.startswith("www."):
                display_host = display_host[4:]

            if (
                site_root == link_host
                or site_root == display_host
                or site_root in link
                or site_root in display_link
            ):
                position = index
                break

    total = data.get("searchInformation", {}).get("totalResults")
    total_int = parse_total_results(total)

    info = {
        "position": position,
        "total_results": total_int,
    }

    cse_brand_cache[cache_key] = info
    return info


def kg_search(name: str) -> Optional[Dict]:
    global kg_requests

    if not GOOGLE_CLOUD_API_KEY or not USE_KNOWLEDGE_GRAPH_API:
        return None
    if not name:
        return None
    if kg_requests >= MAX_KG_REQUESTS:
        return None
    if name in kg_cache:
        return kg_cache[name]

    url = "https://kgsearch.googleapis.com/v1/entities:search"
    params = {
        "key": GOOGLE_CLOUD_API_KEY,
        "query": name,
        "limit": 1,
        "languages": "pt-BR",
        "types": "Organization,LocalBusiness",
    }

    try:
        response = requests.get(url, params=params, timeout=20)
        if response.status_code != 200:
            info = {"has_entity": False, "description": "", "types": ""}
            kg_cache[name] = info
            kg_requests += 1
            return info
        data = response.json()
    except Exception:
        info = {"has_entity": False, "description": "", "types": ""}
        kg_cache[name] = info
        kg_requests += 1
        return info

    item_list = data.get("itemListElement", [])
    if not item_list:
        info = {"has_entity": False, "description": "", "types": ""}
        kg_cache[name] = info
        kg_requests += 1
        return info

    result = item_list[0].get("result", {})
    description = result.get("description") or ""
    types_value = result.get("@type") or []

    if isinstance(types_value, str):
        types_str = types_value
    elif isinstance(types_value, list):
        types_str = ", ".join(types_value)
    else:
        types_str = ""

    info = {
        "has_entity": True,
        "description": description,
        "types": types_str,
    }

    kg_cache[name] = info
    kg_requests += 1
    return info


def compute_scores(row: Dict) -> None:
    rating = row.get("nota_google")
    reviews = row.get("numero_avaliacoes")

    local_score = 0
    if isinstance(rating, (int, float)):
        if rating < 3.5:
            local_score += 40
        elif rating < 4.0:
            local_score += 25
        elif rating < 4.5:
            local_score += 15
        else:
            local_score += 5

    if isinstance(reviews, int):
        if reviews < 5:
            local_score += 30
        elif reviews < 20:
            local_score += 20
        elif reviews < 50:
            local_score += 10

    status_value = row.get("status_google") or ""
    if status_value and status_value != "OPERATIONAL":
        local_score += 20

    if local_score > 100:
        local_score = 100

    ps_performance = row.get("ps_performance_mobile")
    ps_seo = row.get("ps_seo_mobile")
    lcp = row.get("lcp_ms")
    cls = row.get("cls")
    inp = row.get("inp_ms")

    technical_score = 0
    if isinstance(ps_performance, (int, float)):
        technical_score += int((100 - ps_performance) * 0.5)
    if isinstance(ps_seo, (int, float)):
        technical_score += int((100 - ps_seo) * 0.5)

    if isinstance(lcp, (int, float)):
        if lcp > 4000:
            technical_score += 20
        elif lcp > 2500:
            technical_score += 10

    if isinstance(cls, (int, float)):
        if cls > 0.25:
            technical_score += 10
        elif cls > 0.1:
            technical_score += 5

    if isinstance(inp, (int, float)):
        if inp > 400:
            technical_score += 10
        elif inp > 200:
            technical_score += 5

    if technical_score > 100:
        technical_score = 100

    brand_position = row.get("posicao_busca_marca")
    indexed_pages = row.get("paginas_indexadas_aproximadas")
    has_entity = row.get("tem_entidade_kg")

    brand_score = 0
    if isinstance(brand_position, int):
        if brand_position == 1:
            brand_score += 0
        elif brand_position <= 3:
            brand_score += 10
        elif brand_position <= 10:
            brand_score += 25
        else:
            brand_score += 40
    else:
        brand_score += 40

    if isinstance(indexed_pages, int):
        if indexed_pages < 10:
            brand_score += 15
        elif indexed_pages < 50:
            brand_score += 8

    if has_entity is False and isinstance(reviews, int) and reviews > 50:
        brand_score += 20

    if brand_score > 100:
        brand_score = 100

    row["score_local"] = local_score
    row["score_tecnico_seo"] = technical_score
    row["score_visibilidade_marca"] = brand_score

    priority_score = int(round(local_score * 0.4 + technical_score * 0.4 + brand_score * 0.2))
    if priority_score > 100:
        priority_score = 100

    row["score_prioridade_lead"] = priority_score


def save_to_xlsx(path: str, rows: List[Dict]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sanitize_sheet_title(BASE_QUERY)

    headers = [
        "Nome",
        "Cidade",
        "Estado",
        "Endereço",
        "Site",
        "Erro no Site",
        "Telefone / WhatsApp (Google)",
        "Nota no Google",
        "Número de Avaliações",
        "Status no Google",
        "Tipos de Empresa",
        "HTTPS Ativo",
        "Tem Blog",
        "Tem AMP",
        "WhatsApp (site)",
        "Facebook",
        "Instagram",
        "YouTube",
        "TikTok",
        "LinkedIn",
        "Kwai",
        "Messenger",
        "Telegram",
        "Pinterest",
        "X / Twitter",
        "PageSpeed Performance (mobile)",
        "PageSpeed SEO (mobile)",
        "LCP (ms)",
        "CLS",
        "INP (ms)",
        "Categoria de Experiência",
        "Posição na Busca de Marca",
        "Páginas Indexadas (aprox.)",
        "Tem Entidade no KG",
        "Descrição no KG",
        "Tipo de Entidade no KG",
        "Score Local",
        "Score Técnico de SEO",
        "Score de Visibilidade de Marca",
        "Score de Prioridade do Lead",
    ]

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)

    max_length = [len(header) if header else 0 for header in headers]

    for row_index, row in enumerate(rows, start=2):
        name_value = row.get("nome") or ""
        city_value = row.get("cidade") or ""
        state_value = row.get("estado") or ""
        address_value = row.get("endereco") or ""
        maps_url = row.get("google_url") or ""
        website = row.get("website") or ""
        site_error = row.get("site_error") or ""
        phone_value = row.get("telefone") or ""
        whatsapp_phone_link = row.get("whatsapp_telefone_link") or ""
        whatsapp_site_link = row.get("whatsapp_site_link") or ""
        rating_value = row.get("nota_google")
        review_count = row.get("numero_avaliacoes")
        google_status = row.get("status_google") or ""
        business_types = row.get("tipos_empresa") or ""
        https_ok = row.get("https_ok")
        has_blog_value = row.get("tem_blog")
        blog_url = row.get("blog_url")
        has_amp_value = row.get("tem_amp")
        amp_url = row.get("amp_url")
        facebook = row.get("facebook") or ""
        instagram = row.get("instagram") or ""
        youtube = row.get("youtube") or ""
        tiktok = row.get("tiktok") or ""
        linkedin = row.get("linkedin") or ""
        kwai = row.get("kwai") or ""
        messenger = row.get("messenger") or ""
        telegram = row.get("telegram") or ""
        pinterest = row.get("pinterest") or ""
        x_twitter = row.get("x_twitter") or ""
        ps_performance = row.get("ps_performance_mobile")
        ps_seo = row.get("ps_seo_mobile")
        lcp = row.get("lcp_ms")
        cls = row.get("cls")
        inp = row.get("inp_ms")
        experience_category = row.get("experience_category") or ""
        brand_position = row.get("posicao_busca_marca")
        indexed_pages = row.get("paginas_indexadas_aproximadas")
        has_entity = row.get("tem_entidade_kg")
        kg_description = row.get("kg_descricao") or ""
        kg_entity_type = row.get("kg_tipo_entidade") or ""
        local_score = row.get("score_local")
        technical_score = row.get("score_tecnico_seo")
        brand_score = row.get("score_visibilidade_marca")
        priority_score = row.get("score_prioridade_lead")

        column = 1

        if maps_url and name_value:
            formula = f'=HYPERLINK("{maps_url}", "{name_value}")'
            worksheet.cell(row=row_index, column=column, value=formula)
        else:
            worksheet.cell(row=row_index, column=column, value=name_value)
        max_length[column - 1] = max(max_length[column - 1], len(name_value))
        column += 1

        worksheet.cell(row=row_index, column=column, value=city_value)
        max_length[column - 1] = max(max_length[column - 1], len(city_value))
        column += 1

        worksheet.cell(row=row_index, column=column, value=state_value)
        max_length[column - 1] = max(max_length[column - 1], len(state_value))
        column += 1

        if maps_url and address_value:
            formula = f'=HYPERLINK("{maps_url}", "{address_value}")'
            worksheet.cell(row=row_index, column=column, value=formula)
        else:
            worksheet.cell(row=row_index, column=column, value=address_value)
        max_length[column - 1] = max(max_length[column - 1], len(address_value))
        column += 1

        if website:
            worksheet.cell(row=row_index, column=column, value=f'=HYPERLINK("{website}", "{website}")')
        else:
            worksheet.cell(row=row_index, column=column, value=website)
        max_length[column - 1] = max(max_length[column - 1], len(website))
        column += 1

        worksheet.cell(row=row_index, column=column, value=site_error)
        max_length[column - 1] = max(max_length[column - 1], len(site_error))
        column += 1

        phone_display = phone_value or ""
        if whatsapp_phone_link:
            worksheet.cell(
                row=row_index,
                column=column,
                value=f'=HYPERLINK("{whatsapp_phone_link}", "{phone_display}")',
            )
        else:
            worksheet.cell(row=row_index, column=column, value=phone_display)
        max_length[column - 1] = max(max_length[column - 1], len(phone_display))
        column += 1

        if isinstance(rating_value, (int, float)):
            worksheet.cell(row=row_index, column=column, value=rating_value)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(rating_value)) if rating_value is not None else 0,
        )
        column += 1

        if isinstance(review_count, int):
            worksheet.cell(row=row_index, column=column, value=review_count)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(review_count)) if review_count is not None else 0,
        )
        column += 1

        worksheet.cell(row=row_index, column=column, value=google_status)
        max_length[column - 1] = max(max_length[column - 1], len(google_status))
        column += 1

        worksheet.cell(row=row_index, column=column, value=business_types)
        max_length[column - 1] = max(max_length[column - 1], len(business_types))
        column += 1

        https_text = "Sim" if https_ok else "Não" if https_ok is not None else ""
        worksheet.cell(row=row_index, column=column, value=https_text)
        max_length[column - 1] = max(max_length[column - 1], len(https_text))
        column += 1

        blog_text = ""
        if has_blog_value is True:
            blog_text = "Sim"
            if blog_url:
                worksheet.cell(row=row_index, column=column, value=f'=HYPERLINK("{blog_url}", "Sim")')
            else:
                worksheet.cell(row=row_index, column=column, value="Sim")
        elif has_blog_value is False:
            blog_text = "Não"
            worksheet.cell(row=row_index, column=column, value="Não")
        else:
            worksheet.cell(row=row_index, column=column, value="")
        max_length[column - 1] = max(max_length[column - 1], len(blog_text))
        column += 1

        amp_text = ""
        if has_amp_value is True:
            amp_text = "Sim"
            if amp_url:
                worksheet.cell(row=row_index, column=column, value=f'=HYPERLINK("{amp_url}", "Sim")')
            else:
                worksheet.cell(row=row_index, column=column, value="Sim")
        elif has_amp_value is False:
            amp_text = "Não"
            worksheet.cell(row=row_index, column=column, value="Não")
        else:
            worksheet.cell(row=row_index, column=column, value="")
        max_length[column - 1] = max(max_length[column - 1], len(amp_text))
        column += 1

        if whatsapp_site_link:
            worksheet.cell(
                row=row_index,
                column=column,
                value=f'=HYPERLINK("{whatsapp_site_link}", "{whatsapp_site_link}")',
            )
            max_length[column - 1] = max(max_length[column - 1], len(whatsapp_site_link))
        column += 1

        social_links = [
            facebook,
            instagram,
            youtube,
            tiktok,
            linkedin,
            kwai,
            messenger,
            telegram,
            pinterest,
            x_twitter,
        ]
        for social in social_links:
            if social:
                worksheet.cell(
                    row=row_index,
                    column=column,
                    value=f'=HYPERLINK("{social}", "{social}")',
                )
                max_length[column - 1] = max(max_length[column - 1], len(social))
            column += 1

        if isinstance(ps_performance, (int, float)):
            worksheet.cell(row=row_index, column=column, value=ps_performance)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(ps_performance)) if ps_performance is not None else 0,
        )
        column += 1

        if isinstance(ps_seo, (int, float)):
            worksheet.cell(row=row_index, column=column, value=ps_seo)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(ps_seo)) if ps_seo is not None else 0,
        )
        column += 1

        if isinstance(lcp, (int, float)):
            worksheet.cell(row=row_index, column=column, value=lcp)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(lcp)) if lcp is not None else 0,
        )
        column += 1

        if isinstance(cls, (int, float)):
            worksheet.cell(row=row_index, column=column, value=cls)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(cls)) if cls is not None else 0,
        )
        column += 1

        if isinstance(inp, (int, float)):
            worksheet.cell(row=row_index, column=column, value=inp)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(inp)) if inp is not None else 0,
        )
        column += 1

        worksheet.cell(row=row_index, column=column, value=experience_category)
        max_length[column - 1] = max(max_length[column - 1], len(experience_category))
        column += 1

        if isinstance(brand_position, int):
            worksheet.cell(row=row_index, column=column, value=brand_position)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(brand_position)) if brand_position is not None else 0,
        )
        column += 1

        if isinstance(indexed_pages, int):
            worksheet.cell(row=row_index, column=column, value=indexed_pages)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(indexed_pages)) if indexed_pages is not None else 0,
        )
        column += 1

        if has_entity is True:
            entity_text = "Sim"
        elif has_entity is False:
            entity_text = "Não"
        else:
            entity_text = ""
        worksheet.cell(row=row_index, column=column, value=entity_text)
        max_length[column - 1] = max(max_length[column - 1], len(entity_text))
        column += 1

        worksheet.cell(row=row_index, column=column, value=kg_description)
        max_length[column - 1] = max(max_length[column - 1], len(kg_description))
        column += 1

        worksheet.cell(row=row_index, column=column, value=kg_entity_type)
        max_length[column - 1] = max(max_length[column - 1], len(kg_entity_type))
        column += 1

        if isinstance(local_score, (int, float)):
            worksheet.cell(row=row_index, column=column, value=local_score)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(local_score)) if local_score is not None else 0,
        )
        column += 1

        if isinstance(technical_score, (int, float)):
            worksheet.cell(row=row_index, column=column, value=technical_score)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(technical_score)) if technical_score is not None else 0,
        )
        column += 1

        if isinstance(brand_score, (int, float)):
            worksheet.cell(row=row_index, column=column, value=brand_score)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(brand_score)) if brand_score is not None else 0,
        )
        column += 1

        if isinstance(priority_score, (int, float)):
            worksheet.cell(row=row_index, column=column, value=priority_score)
        max_length[column - 1] = max(
            max_length[column - 1],
            len(str(priority_score)) if priority_score is not None else 0,
        )

    for index, length in enumerate(max_length, start=1):
        width = min(length + 2, 60)
        column_letter = get_column_letter(index)
        worksheet.column_dimensions[column_letter].width = width

    workbook.save(path)


def collect_companies() -> None:
    global pagespeed_cache, cse_site_cache, cse_brand_cache, kg_cache

    if not GOOGLE_CLOUD_API_KEY:
        raise RuntimeError("Defina GOOGLE_CLOUD_API_KEY no arquivo .env")
    if not USE_PLACES_API:
        logging.error("USE_PLACES_API está desativado. Nenhuma empresa será coletada.")
        return

    pagespeed_cache = load_json_cache(PAGESPEED_CACHE_FILE)
    cse_site_cache = load_json_cache(CSE_SITE_CACHE_FILE)
    cse_brand_cache = load_json_cache(CSE_BRAND_CACHE_FILE)
    kg_cache = load_json_cache(KG_CACHE_FILE)

    known_place_ids = load_known_places()
    all_rows: List[Dict] = []

    for city_query in CITIES:
        logging.info(f"Iniciando busca por '{BASE_QUERY}' em {city_query}")
        query = f"{BASE_QUERY} {city_query}"
        page_token = None

        while True:
            try:
                search_data = google_places_text_search(query, GOOGLE_CLOUD_API_KEY, page_token=page_token)
            except Exception as exception:
                logging.error(f"Erro na API Places para {city_query}: {exception}")
                break

            results = search_data.get("places", [])
            if not results:
                break

            for place in results:
                place_id = place.get("id")
                if not place_id:
                    continue
                if place_id in known_place_ids:
                    continue

                known_place_ids.add(place_id)

                try:
                    details = google_place_details(place_id, GOOGLE_CLOUD_API_KEY)
                except Exception as exception:
                    logging.warning(f"Erro ao obter detalhes de {place_id}: {exception}")
                    continue

                display_name = details.get("displayName") or {}
                if isinstance(display_name, dict):
                    name_value = display_name.get("text")
                else:
                    name_value = None

                address_value = details.get("formattedAddress")
                website = details.get("websiteUri")
                phone_international = details.get("internationalPhoneNumber")
                phone_national = details.get("nationalPhoneNumber")
                phone_to_use = phone_international or phone_national
                rating_value = details.get("rating")
                review_count = details.get("userRatingCount")
                types_list = details.get("types") or []
                google_status = details.get("businessStatus")
                maps_url = details.get("googleMapsUri") or f"https://www.google.com/maps/place/?q=place_id:{place_id}"

                if not website or not phone_to_use:
                    continue

                site_type, has_real_site = classify_website(website)
                if not has_real_site:
                    continue

                https_ok = None
                has_blog_value = None
                blog_url = None
                has_amp_value = None
                amp_url = None
                facebook = None
                instagram = None
                youtube = None
                tiktok = None
                linkedin = None
                kwai = None
                messenger = None
                telegram = None
                pinterest = None
                x_twitter = None
                has_whatsapp = False
                whatsapp_site_link = None
                site_error = None

                if has_real_site:
                    site_info = site_scrape(website)
                    https_ok = site_info.get("https_ok")
                    has_blog_value = site_info.get("tem_blog")
                    blog_url = site_info.get("blog_url")
                    has_amp_value = site_info.get("tem_amp")
                    amp_url = site_info.get("amp_url")
                    facebook = site_info.get("facebook")
                    instagram = site_info.get("instagram")
                    youtube = site_info.get("youtube")
                    tiktok = site_info.get("tiktok")
                    linkedin = site_info.get("linkedin")
                    kwai = site_info.get("kwai")
                    messenger = site_info.get("messenger")
                    telegram = site_info.get("telegram")
                    pinterest = site_info.get("pinterest")
                    x_twitter = site_info.get("x_twitter")
                    has_whatsapp = site_info.get("has_whatsapp") or False
                    whatsapp_site_link = site_info.get("whatsapp_url")
                    site_error = site_info.get("site_error")

                if not has_whatsapp and site_error is None:
                    continue

                final_city, final_state = extract_city_state(address_value, search_city=city_query)
                whatsapp_phone_link = normalize_phone_to_whatsapp_link(phone_to_use)
                domain = get_domain_root(website) if has_real_site else None

                row_data = {
                    "nome": name_value or "",
                    "cidade": final_city or "",
                    "estado": final_state or "",
                    "endereco": address_value or "",
                    "google_url": maps_url,
                    "website": website,
                    "site_error": site_error,
                    "telefone": phone_to_use,
                    "whatsapp_telefone_link": whatsapp_phone_link,
                    "whatsapp_site_link": whatsapp_site_link,
                    "nota_google": rating_value,
                    "numero_avaliacoes": review_count if isinstance(review_count, int) else None,
                    "status_google": google_status or "",
                    "tipos_empresa": ", ".join(types_list) if types_list else "",
                    "https_ok": https_ok,
                    "tem_blog": has_blog_value,
                    "blog_url": blog_url,
                    "tem_amp": has_amp_value,
                    "amp_url": amp_url,
                    "facebook": facebook,
                    "instagram": instagram,
                    "youtube": youtube,
                    "tiktok": tiktok,
                    "linkedin": linkedin,
                    "kwai": kwai,
                    "messenger": messenger,
                    "telegram": telegram,
                    "pinterest": pinterest,
                    "x_twitter": x_twitter,
                    "dominio": domain,
                    "ps_performance_mobile": None,
                    "ps_seo_mobile": None,
                    "lcp_ms": None,
                    "cls": None,
                    "inp_ms": None,
                    "experience_category": None,
                    "posicao_busca_marca": None,
                    "paginas_indexadas_aproximadas": None,
                    "tem_entidade_kg": None,
                    "kg_descricao": None,
                    "kg_tipo_entidade": None,
                    "score_local": None,
                    "score_tecnico_seo": None,
                    "score_visibilidade_marca": None,
                    "score_prioridade_lead": None,
                }

                all_rows.append(row_data)

            page_token = search_data.get("nextPageToken")
            if not page_token:
                break
            time.sleep(2)

    logging.info(f"Coleta de Places finalizada. Empresas coletadas antes das análises: {len(all_rows)}")

    unique_domains = sorted({row["dominio"] for row in all_rows if row.get("dominio")})

    if unique_domains:
        logging.info(f"Iniciando PageSpeed Insights para {len(unique_domains)} domínios")

    for index, domain in enumerate(unique_domains, start=1):
        domain_rows = [row for row in all_rows if row.get("dominio") == domain]
        if not domain_rows:
            continue

        site_url = domain_rows[0].get("website")
        if not site_url:
            continue

        logging.info(f"PageSpeed {index}/{len(unique_domains)} para {domain}")
        pagespeed_data = get_pagespeed_for_domain(domain, site_url)
        if not pagespeed_data:
            continue

        for row in domain_rows:
            row["ps_performance_mobile"] = pagespeed_data.get("ps_performance")
            row["ps_seo_mobile"] = pagespeed_data.get("ps_seo")
            row["lcp_ms"] = pagespeed_data.get("lcp_ms")
            row["cls"] = pagespeed_data.get("cls")
            row["inp_ms"] = pagespeed_data.get("inp_ms")
            row["experience_category"] = pagespeed_data.get("experience_category")

    if USE_PROGRAMMABLE_SEARCH_ENGINE_API and GOOGLE_PROGRAMMABLE_SEARCH_ENGINE_CX:
        if unique_domains:
            logging.info("Iniciando análise de indexação (CSE)")

        for domain in unique_domains:
            site_index_info = get_site_index_info(domain)
            total_indexed = site_index_info.get("total_results") if site_index_info else None
            for row in all_rows:
                if row.get("dominio") == domain:
                    row["paginas_indexadas_aproximadas"] = total_indexed

        logging.info("Iniciando análise de busca de marca (CSE)")
        for row in all_rows:
            name_value = row.get("nome") or ""
            city_value = row.get("cidade") or ""
            domain = row.get("dominio")
            brand_info = get_brand_search_info(domain, row.get("website"), name_value, city_value)
            if brand_info:
                row["posicao_busca_marca"] = brand_info.get("position")
    else:
        logging.info("Programmable Search Engine desabilitado (sem CX configurado ou uso desativado)")

    if USE_KNOWLEDGE_GRAPH_API:
        logging.info("Iniciando consultas ao Knowledge Graph")

    for index, row in enumerate(all_rows, start=1):
        name_value = row.get("nome") or ""
        if not name_value:
            row["tem_entidade_kg"] = None
            continue

        kg_info = kg_search(name_value)
        if not kg_info:
            row["tem_entidade_kg"] = False
            continue

        row["tem_entidade_kg"] = kg_info.get("has_entity", False)
        row["kg_descricao"] = kg_info.get("description")
        row["kg_tipo_entidade"] = kg_info.get("types")

        if USE_KNOWLEDGE_GRAPH_API and index % 10 == 0:
            logging.info(f"Knowledge Graph: {index}/{len(all_rows)} empresas processadas")

    logging.info("Calculando scores de prioridade")
    for row in all_rows:
        compute_scores(row)

    logging.info(f"Salvando planilha em {OUTPUT_XLSX}")
    save_to_xlsx(OUTPUT_XLSX, all_rows)

    logging.info("Salvando caches em disco")
    save_known_places(known_place_ids)
    save_json_cache(PAGESPEED_CACHE_FILE, pagespeed_cache)
    save_json_cache(CSE_SITE_CACHE_FILE, cse_site_cache)
    save_json_cache(CSE_BRAND_CACHE_FILE, cse_brand_cache)
    save_json_cache(KG_CACHE_FILE, kg_cache)

    logging.info(f"Coleta finalizada. Total de empresas salvas: {len(all_rows)}")


if __name__ == "__main__":
    collect_companies()
